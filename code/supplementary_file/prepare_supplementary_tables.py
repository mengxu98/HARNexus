#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from functions.utils import log_message

PROJECT_ROOT = os.path.dirname(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
)
OUT_DIR = os.path.join(PROJECT_ROOT, "results", "supplementary_file")
XLSX_PATH = os.path.join(OUT_DIR, "Supplementary Tables.xlsx")
FONT_NAME = "Times New Roman"
HEADER_FILL = "D3D3D3"  # light gray

_PRESERVE_WORDS_UPPER = {"TFS", "HARS", "HAR_ID", "HIC"}
SINGULAR_TO_PLURAL = {"HAR": "HARs", "TF": "TFs", "Gene": "Genes", "Motif": "Motifs"}
_SINGULAR_LOWER = {"har": "HARs", "tf": "TFs", "gene": "Genes", "motif": "Motifs"}
_KEEP_AS_IS_UPPER = {"HARS", "HAR_ID", "TFS"}
_KEEP_CANONICAL = {"HARS": "HARs", "HAR_ID": "HAR_ID", "TFS": "TFs"}


def _format_column_name(name):
    key = name.strip()
    if key in SINGULAR_TO_PLURAL:
        return SINGULAR_TO_PLURAL[key]
    if key.lower() in _SINGULAR_LOWER:
        return _SINGULAR_LOWER[key.lower()]
    if key.upper() in _KEEP_AS_IS_UPPER:
        return _KEEP_CANONICAL.get(key.upper(), key)
    s = name.replace("_", " ")
    s_lower = s.lower().strip()
    if s_lower in ("celltype", "celltypes", "cell type"):
        return "Cell types"
    s = s.replace("BrainRegion", "Brain region")
    if s.startswith("n ") and len(s) > 2:
        s = "Number of " + s[2:]
    elif len(s) > 1 and s[0] == "n" and s[1].isupper():
        s = "Number of " + s[1:]
    s = re.sub(r"\bChimp\b", "chimpanzee", s, flags=re.IGNORECASE)
    s = re.sub(r"\bhic\b", "HiC", s, flags=re.IGNORECASE)
    parts = s.split()
    if not parts:
        return s
    rest = [w if w.upper() in _PRESERVE_WORDS_UPPER else w.lower() for w in parts[1:]]
    result = " ".join([parts[0].capitalize()] + rest)
    if result == "Direction":
        return "Directions"
    if "hi-c" in result.lower() and "validated target genes" in result.lower():
        result = re.sub(r"\bhi-c\b", "Hi-C", result, flags=re.IGNORECASE)
    return result


def _format_df_columns(df):
    if df.empty:
        return df
    df = df.copy()
    df.columns = [_format_column_name(c) for c in df.columns]
    return df


def _path(*parts):
    return os.path.join(PROJECT_ROOT, *parts)


def load_sheet1():
    df_h = pd.read_csv(_path("results", "har_tf", "human", "har_tf_pairs.csv"))
    df_c = pd.read_csv(_path("results", "har_tf", "chimp", "har_tf_pairs.csv"))
    if "Species" not in df_h.columns:
        df_h["Species"] = "Human"
    if "Species" not in df_c.columns:
        df_c["Species"] = "Chimpanzee"
    df = pd.concat([df_h, df_c], ignore_index=True)
    df = df.rename(
        columns={
            "har": "HARs",
            "start": "Start",
            "end": "End",
            "strand": "Strand",
            "motif": "Motifs",
            "TF": "TFs",
            "score": "Score",
        }
    )
    df = df[["HARs", "Start", "End", "Strand", "Motifs", "TFs", "Score", "Species"]]
    return df.sort_values("Score", ascending=False).reset_index(drop=True)


def load_sheet2():
    df = pd.read_csv(_path("results", "har_tf", "motif_score_aggregated.csv"))
    df = df.rename(
        columns={
            "HAR_ID": "HAR",
            "TF_Name": "TF",
            "Mean_Human_Score": "Human score",
            "Mean_Chimp_Score": "Chimp score",
            "Num_Motifs": "Number of motifs",
            "Mean_Diff": "Score difference",
            "Direction": "Directions",
        }
    )
    return df.sort_values("Score difference", ascending=False).reset_index(drop=True)


def load_sheet3():
    p = _path("results", "hic", "HAR_gene_HiC_supported.csv")
    if not os.path.exists(p):
        return pd.DataFrame()
    df = pd.read_csv(p)
    sort_col = "Number of Hi-C-validated target genes"
    df = df.rename(
        columns={
            "HAR": "HARs",
            "hic_gene": "HiC target genes",
            "n_HiC_validation": sort_col,
        }
    )
    df = df[["HARs", "HAR_ID", "HiC target genes", sort_col]]
    return df.sort_values(sort_col, ascending=False).reset_index(drop=True)


def _load_astro_network(method_key, filename, top_n=500):
    p = _path("results", "gse97942", filename)
    if not os.path.exists(p):
        return pd.DataFrame()
    df = pd.read_csv(p)
    if df.empty or "weight" not in df.columns:
        return df
    df = df.copy()
    df["abs_weight"] = df["weight"].abs()
    df = (
        df.sort_values("abs_weight", ascending=False)
        .drop(columns=["abs_weight"])
        .reset_index(drop=True)
    )
    df = df.head(top_n)
    return df.rename(
        columns={"regulator": "TFs", "target": "Target genes", "weight": "Weights"}
    )


def load_sheet4():
    return _load_astro_network("GENIE3", "genie3_network.csv")


def load_sheet5():
    return _load_astro_network("HARNexus", "infercsn_network.csv")


def load_sheet6():
    return _load_astro_network("LEAP", "leap_network.csv")


def load_sheet7():
    return _load_astro_network("PPCOR", "ppcor_network.csv")


def load_sheet8():
    p = _path("results", "networks", "analysis", "network_data.csv")
    if not os.path.exists(p):
        return pd.DataFrame()
    required = ["Region", "Stage", "CellType", "TF", "Target"]
    usecols = required
    chunks = []
    for chunk in pd.read_csv(p, usecols=usecols, chunksize=500_000):
        chunks.append(chunk)
    df = pd.concat(chunks, ignore_index=True)
    if not all(c in df.columns for c in required):
        return pd.DataFrame()

    def agg_group(g):
        tfs = set(g["TF"])
        targets_non_tf = g[~g["Target"].isin(tfs)]["Target"]
        return pd.Series(
            {
                "Number of edges": len(g),
                "Number of TFs": g["TF"].nunique(),
                "Number of target genes": targets_non_tf.nunique(),
            }
        )

    stats = (
        df.groupby(["Region", "Stage", "CellType"], dropna=False)
        .apply(agg_group)
        .reset_index()
    )
    stats = stats.rename(
        columns={"Region": "Brain regions", "Stage": "Stages", "CellType": "Cell types"}
    )
    return stats[
        [
            "Brain regions",
            "Stages",
            "Cell types",
            "Number of edges",
            "Number of TFs",
            "Number of target genes",
        ]
    ]


def _load_species_networks(pair_id, human_prefix, chimp_prefix):
    csv_dir = _path("results", "species_networks", pair_id, "csv")
    if not os.path.isdir(csv_dir):
        return pd.DataFrame()
    frames = []
    for f in os.listdir(csv_dir):
        if not f.endswith(".csv"):
            continue
        full = os.path.join(csv_dir, f)
        if f.startswith(human_prefix):
            species = "Human"
            celltype = f[len(human_prefix) : -4]
        elif f.startswith(chimp_prefix):
            species = "Chimpanzee"
            celltype = f[len(chimp_prefix) : -4]
        else:
            continue
        try:
            df = pd.read_csv(full)
        except Exception:
            continue
        df = df.rename(
            columns={"regulator": "TFs", "target": "Target genes", "weight": "Weights"}
        )
        df["Species"] = species
        df["CellType"] = celltype
        frames.append(df)
    if not frames:
        return pd.DataFrame()
    out = pd.concat(frames, ignore_index=True)
    cols = ["Species", "CellType"] + [
        c for c in out.columns if c not in ("Species", "CellType")
    ]
    return out[cols]


def load_sheet9():
    df1 = _load_species_networks("h3_c2", "human_h3_", "chimp_c2_")
    df2 = _load_species_networks("h4_c4", "human_h4_", "chimp_c4_")
    if df1.empty and df2.empty:
        return pd.DataFrame()
    if df1.empty:
        df2.insert(0, "Samples", "h4_c4")
        return df2
    if df2.empty:
        df1.insert(0, "Samples", "h3_c2")
        return df1
    df1.insert(0, "Samples", "h3_c2")
    df2.insert(0, "Samples", "h4_c4")
    out = pd.concat([df1, df2], ignore_index=True)
    cols = ["Samples"] + [c for c in out.columns if c != "Samples"]
    return out[cols]


def _load_evolution_one(pair_id):
    p = _path(
        "results", "species_networks", pair_id, "evolution_daccre_for_target_genes.csv"
    )
    if not os.path.exists(p):
        return pd.DataFrame()
    return _rename_evolution_columns(pd.read_csv(p))


def load_sheet10():
    df1 = _load_evolution_one("h3_c2")
    df2 = _load_evolution_one("h4_c4")
    if df1.empty and df2.empty:
        return pd.DataFrame()
    if df1.empty:
        df2.insert(0, "Samples", "h4_c4")
        if "n_peaks" in df2.columns:
            df2 = df2.sort_values("n_peaks", ascending=False).reset_index(drop=True)
        return df2
    if df2.empty:
        df1.insert(0, "Samples", "h3_c2")
        if "n_peaks" in df1.columns:
            df1 = df1.sort_values("n_peaks", ascending=False).reset_index(drop=True)
        return df1
    df1.insert(0, "Samples", "h3_c2")
    df2.insert(0, "Samples", "h4_c4")
    out = pd.concat([df1, df2], ignore_index=True)
    cols = ["Samples"] + [c for c in out.columns if c != "Samples"]
    out = out[cols]
    if "n_peaks" in out.columns:
        out = out.sort_values("n_peaks", ascending=False).reset_index(drop=True)
    return out


def _rename_evolution_columns(df):
    rename = {
        "Gene": "Target genes",
        "Target_type": "Target gene types",
        "Peak_type": "Peak types",
        "Evolution_type": "Evolution types",
    }
    for c in df.columns:
        if c not in rename and c.endswith("type"):
            base = c[:-5].rstrip("_") if c.endswith("_type") else c[:-4]
            rename[c] = base + " types"
    return df.rename(columns={k: v for k, v in rename.items() if k in df.columns})


def load_sheet11():
    p_t = _path("results", "pfc_astrocytes", "target_genes.csv")
    p_v = _path("results", "pfc_astrocytes", "var_genes.csv")
    target_col = "Target_genes"
    var_col = "Variable_genes"
    if os.path.exists(p_t):
        df_t = pd.read_csv(p_t)
        t_genes = df_t.iloc[:, 0].astype(str).tolist()
    else:
        t_genes = []
    if os.path.exists(p_v):
        df_v = pd.read_csv(p_v)
        v_genes = df_v.iloc[:, 0].astype(str).tolist()
    else:
        v_genes = []
    n = max(len(t_genes), len(v_genes))
    if n == 0:
        return pd.DataFrame({target_col: [], var_col: []})
    out = pd.DataFrame(
        {
            target_col: (t_genes + [None] * (n - len(t_genes)))[:n],
            var_col: (v_genes + [None] * (n - len(v_genes)))[:n],
        }
    )
    return out


def load_sheet12():
    p = _path("results", "pfc_astrocytes", "dynamic_heatmap_target_feature.csv")
    if not os.path.exists(p):
        return pd.DataFrame()
    df = pd.read_csv(p)
    df.columns = ["Target genes", "Clusters"]
    return df


def main():
    log_message("Building Supplementary Tables...", message_type="info")
    os.makedirs(OUT_DIR, exist_ok=True)

    loaders = [
        (1, "HAR_TF_motif", load_sheet1),
        (2, "TFs_binding_score", load_sheet2),
        (3, "HiC_validated_target_genes", load_sheet3),
        (4, "GENIE3_astrocytes", load_sheet4),
        (5, "HARNexus_astrocytes", load_sheet5),
        (6, "LEAP_astrocytes", load_sheet6),
        (7, "PPCOR_astrocytes", load_sheet7),
        (8, "HAR_CSN_atlas_stats", load_sheet8),
        (9, "Species_CSNs", load_sheet9),
        (10, "Evolution_of_target_genes", load_sheet10),
        (11, "Prefrontal_cortex_astrocytes", load_sheet11),
        (12, "Dynamic_target_genes", load_sheet12),
    ]

    table_info = pd.DataFrame(
        {
            "Tables": [f"Table {i}" for i in range(1, len(loaders) + 1)],
            "Sheet name": [f"{num}_{sheet_name}" for num, sheet_name, _ in loaders],
        }
    )

    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
        table_info.to_excel(writer, sheet_name="Table information", index=False)
        log_message("Sheet Table information: written.", message_type="info")
        for num, sheet_name, loader in loaders:
            try:
                df = loader()
            except Exception as e:
                log_message(
                    f"Sheet {num} ({sheet_name}): load failed: {e}",
                    message_type="warning",
                )
                df = pd.DataFrame()
            if df.empty:
                df = pd.DataFrame({"Note": ["Data not found or empty."]})
            df = _format_df_columns(df)
            sheet_name_with_num = f"{num}_{sheet_name}"
            csv_path = os.path.join(OUT_DIR, f"sheet{num:02d}_{sheet_name}.csv")
            df.to_csv(csv_path, index=False)
            df.to_excel(writer, sheet_name=sheet_name_with_num[:31], index=False)
            log_message(
                f"Sheet {num:2d} {sheet_name_with_num}: {len(df):,} rows -> {csv_path}",
                message_type="info",
            )

    log_message("Applying styles and column widths...", message_type="running")
    wb = load_workbook(XLSX_PATH)
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
    header_font = Font(name=FONT_NAME, bold=True)
    body_font = Font(name=FONT_NAME)
    gene_font = Font(name=FONT_NAME, italic=True)
    gene_col_names = (
        "TFs",
        "Target genes",
        "HiC target genes",
        "Hic target genes",
        "Variable genes",
    )
    for ws in wb.worksheets:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.font = body_font
        col_names_row1 = [
            ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)
        ]
        for col_idx, name in enumerate(col_names_row1, start=1):
            if name in gene_col_names:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).font = gene_font
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(1, min(ws.max_row + 1, 102)):
                c = ws.cell(row=row_idx, column=col_idx)
                if c.value is not None:
                    max_len = max(max_len, len(str(c.value)))
            w = min(max(max_len + 1, 8), 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = w
    wb.save(XLSX_PATH)
    log_message(f"XLSX saved: {XLSX_PATH}", message_type="success")
    log_message("Supplementary tables build complete", message_type="success")


if __name__ == "__main__":
    main()
